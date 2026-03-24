import streamlit as st
from fpdf import FPDF
import openpyxl
import io
import requests
from bs4 import BeautifulSoup
from groq import Groq
from datetime import date
from PIL import Image
import json
import re

st.title("Analizador de Idealista")

url = st.text_input("Pega la URL del anuncio de Idealista")

if "pdf_bytes" not in st.session_state:
    st.session_state.pdf_bytes = None
if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None

if st.button("Analizar") and url:
    with st.spinner("Analizando el anuncio..."):

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "es-ES,es;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Referer": "https://www.google.es/"
        }

        session = requests.Session()
        session.get("https://www.idealista.com", headers=headers)
        response = session.get(url, headers=headers)
        soup = BeautifulSoup(response.text, "html.parser")

        for tag in soup(["script", "style"]):
            tag.decompose()
        html_limpio = soup.get_text(separator="\n", strip=True)[:8000]


        

        st.text_area("HTML recibido (debug)", html_limpio[:1000])



        

        foto_url = None
        for img in soup.find_all("img"):
            src = img.get("src", "")
            if "images.idealista.com" in src:
                foto_url = src
                break

        client = Groq(api_key=st.secrets["GROQ_API_KEY"])

        prompt = f"""
        Extrae los siguientes datos del anuncio de Idealista.
        Devuelve SOLO un JSON sin texto adicional ni backticks.
        Campos: numero_anuncio, provincia, municipio, codigo_postal,
        direccion, precio, tipo_inmueble, superficie, antiguedad.
        Si no encuentras un dato pon "No disponible".

        Texto del anuncio:
        {html_limpio}
        """

        respuesta = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}]
        )

        texto = respuesta.choices[0].message.content.strip()
        texto = re.sub(r"```json|```", "", texto).strip()
        datos = json.loads(texto)
        datos["fecha_obtencion"] = str(date.today())

        foto_bytes = None
        if foto_url:
            try:
                foto_resp = session.get(foto_url, headers=headers)
                foto_bytes = foto_resp.content
            except:
                pass

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "Ficha del inmueble", ln=True, align="C")
        pdf.ln(5)
        pdf.set_font("Helvetica", size=11)

        campos = {
            "Numero anuncio": datos.get("numero_anuncio", ""),
            "Fecha obtencion": datos.get("fecha_obtencion", ""),
            "Provincia": datos.get("provincia", ""),
            "Municipio": datos.get("municipio", ""),
            "Codigo postal": datos.get("codigo_postal", ""),
            "Direccion": datos.get("direccion", ""),
            "Precio": datos.get("precio", ""),
            "Tipo inmueble": datos.get("tipo_inmueble", ""),
            "Superficie": datos.get("superficie", ""),
            "Antiguedad": datos.get("antiguedad", ""),
        }

        for clave, valor in campos.items():
            pdf.cell(0, 9, f"{clave}: {valor}", ln=True)

        if foto_bytes:
            try:
                img = Image.open(io.BytesIO(foto_bytes))
                img_path = "/tmp/foto.jpg"
                img.save(img_path)
                pdf.ln(5)
                pdf.image(img_path, w=150)
            except:
                pass

        st.session_state.pdf_bytes = bytes(pdf.output())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inmueble"
        ws.append(list(campos.keys()))
        ws.append(list(campos.values()))
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        st.session_state.excel_buffer = excel_buffer.read()

        st.success("Archivos generados correctamente")

if st.session_state.pdf_bytes and st.session_state.excel_buffer:
    col1, col2 = st.columns(2)
    with col1:
        st.download_button("Descargar PDF", data=st.session_state.pdf_bytes,
            file_name="inmueble.pdf", mime="application/pdf")
    with col2:
        st.download_button("Descargar Excel", data=st.session_state.excel_buffer,
            file_name="inmueble.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
